"""
读取私有费效分析 Excel，并只导入费效分析相关 4 个 sheet 的汇总结果。

用法:
  python scripts/import_fee_effect.py --dry-run
  python scripts/import_fee_effect.py --confirm
  python scripts/import_fee_effect.py --input private-data/费效分析0414.xlsx --confirm
"""

from __future__ import annotations

import argparse
import hashlib
import os
import re
import sys
import time
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable

import httpx
import openpyxl

ROOT_DIR = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = ROOT_DIR / "private-data" / "费效分析0414.xlsx"
BATCH_TABLE = "fee_effect_import_batches"
BATCH_SIZE = 500
REQUEST_RETRIES = 3

PERSON_SUMMARY_SHEET = "费效分析—1.1 人员维度费效分析"
PERSON_TRAVEL_SHEET = "费效分析—1.2 人员维度差旅明细"
PERSON_HOSPITALITY_SHEET = "费效分析—1.3 人员维度招待明细"
PROJECT_SUMMARY_SHEET = "费效分析—2.项目维度费效分析汇总"

SHEET_ALIASES = {
    PERSON_SUMMARY_SHEET: [
        "费效分析—1.1 人员维度费效分析",
        "费效分析—1.1 人员维度费效分析（汇总）",
    ],
    PERSON_TRAVEL_SHEET: [PERSON_TRAVEL_SHEET],
    PERSON_HOSPITALITY_SHEET: [PERSON_HOSPITALITY_SHEET],
    PROJECT_SUMMARY_SHEET: [PROJECT_SUMMARY_SHEET],
}

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="导入费效分析 4 个汇总 sheet 到 Supabase")
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="费效分析 Excel 文件路径")
    parser.add_argument("--dry-run", action="store_true", help="仅解析并输出摘要，不写入数据库")
    parser.add_argument("--confirm", action="store_true", help="确认执行覆盖写入")
    parser.add_argument("--no-clear", action="store_true", help="不清空旧批次，仅追加新批次")
    return parser.parse_args()


def load_app_env() -> None:
    env_path = ROOT_DIR / "app" / ".env"
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, value = line.split("=", 1)
            os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def require_supabase_env() -> tuple[str, str, str]:
    load_app_env()
    supabase_url = os.getenv("SUPABASE_URL") or os.getenv("VITE_SUPABASE_URL")
    write_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("VITE_SUPABASE_ANON_KEY")
    auth_mode = "service_role" if os.getenv("SUPABASE_SERVICE_ROLE_KEY") else "anon"
    if not supabase_url or not write_key:
        print("错误: 缺少 SUPABASE_URL/VITE_SUPABASE_URL 或 SUPABASE_SERVICE_ROLE_KEY/VITE_SUPABASE_ANON_KEY")
        sys.exit(1)
    return supabase_url, write_key, auth_mode


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\r", "").strip()
    return text or None


def to_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text or text in {"-", "—"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def to_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    return datetime.fromisoformat(text.replace("/", "-")).date().isoformat()


def cell(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def rows(ws: Any, min_row: int) -> list[tuple[Any, ...]]:
    if hasattr(ws, "reset_dimensions"):
        ws.reset_dimensions()
    return list(ws.iter_rows(min_row=min_row, values_only=True))


def normalize_header(value: Any) -> str:
    text = clean_text(value) or ""
    return re.sub(r"\s+", "", text)


def resolve_sheet(wb: Any, canonical_name: str) -> Any:
    aliases = SHEET_ALIASES.get(canonical_name, [canonical_name])
    for name in aliases:
        if name in wb.sheetnames:
            return wb[name]
    for ws in wb.worksheets:
        normalized_title = normalize_header(ws.title)
        if any(normalize_header(alias) in normalized_title for alias in aliases):
            return ws
    expected = " / ".join(aliases)
    raise KeyError(f"找不到 sheet: {expected}")


def find_header_row(ws: Any, required_headers: list[str], max_row: int = 12) -> tuple[int, dict[str, int]]:
    required = {normalize_header(header) for header in required_headers}
    for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, values_only=True), 1):
        header_index = {
            normalized: index
            for index, value in enumerate(row)
            if (normalized := normalize_header(value))
        }
        if required.issubset(header_index.keys()):
            return row_index, header_index
    raise ValueError(f"{ws.title} 未找到表头，缺少: {', '.join(required_headers)}")


def value_by_header(row: tuple[Any, ...], header_index: dict[str, int], *headers: str) -> Any:
    for header in headers:
        index = header_index.get(normalize_header(header))
        if index is not None:
            return cell(row, index)
    return None


def number_by_header(row: tuple[Any, ...], header_index: dict[str, int], *headers: str) -> float:
    return to_number(value_by_header(row, header_index, *headers)) or 0.0


def period_range(wb: Any) -> tuple[str, str]:
    text = clean_text(resolve_sheet(wb, PERSON_SUMMARY_SHEET)["A1"].value) or ""
    match = re.search(r"(\d{6})\s*-\s*(\d{6})", text)
    return (match.group(1), match.group(2)) if match else ("", "")


def parse_person_summary_rows(wb: Any) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    ws = resolve_sheet(wb, PERSON_SUMMARY_SHEET)
    header_row, header_index = find_header_row(ws, ["所属部门", "姓名"])
    for row in rows(ws, header_row + 1):
        department = clean_text(value_by_header(row, header_index, "所属部门"))
        person_name = clean_text(value_by_header(row, header_index, "姓名", "人员"))
        if not person_name or "合计" in person_name:
            continue
        travel_transportation = number_by_header(row, header_index, "差旅-交通（万元）", "差旅-交通")
        travel_lodging = number_by_header(row, header_index, "差旅-住宿（万元）", "差旅-住宿")
        travel_allowance = number_by_header(row, header_index, "差旅-差补（万元）", "差旅-差补")
        travel_total = number_by_header(row, header_index, "差旅合计（万元）", "差旅合计")
        if not travel_total:
            travel_total = travel_transportation + travel_lodging + travel_allowance
        hospitality_amount = number_by_header(row, header_index, "业务招待合计（万元）", "业务招待费（万元）", "业务招待合计")
        total_expense = number_by_header(row, header_index, "差旅招待总计（万元）", "差旅招待总计", "费用合计")
        if not total_expense:
            total_expense = travel_total + hospitality_amount
        records.append({
            "department": department,
            "person_name": person_name,
            "signing_revenue_amount": number_by_header(row, header_index, "签单营收金额（万元）", "签单营收金额"),
            "signing_profit_amount": number_by_header(row, header_index, "签单利润金额（万元）", "签单利润金额"),
            "travel_transportation_amount": travel_transportation,
            "travel_lodging_amount": travel_lodging,
            "travel_allowance_amount": travel_allowance,
            "travel_total_amount": travel_total,
            "hospitality_total_amount": hospitality_amount,
            "total_expense_amount": total_expense,
        })
    return records


def parse_person_travel_rows(wb: Any) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    ws = resolve_sheet(wb, PERSON_TRAVEL_SHEET)
    header_row, header_index = find_header_row(ws, ["人员", "MDM项目名称"])
    for row in rows(ws, header_row + 1):
        person_name = clean_text(value_by_header(row, header_index, "人员", "姓名"))
        department = clean_text(value_by_header(row, header_index, "当前所属部门", "所属部门"))
        mdm_project_name = clean_text(value_by_header(row, header_index, "MDM项目名称"))
        if not person_name or not mdm_project_name:
            continue
        travel_transportation = number_by_header(row, header_index, "差旅-交通（万元）", "差旅-交通")
        travel_lodging = number_by_header(row, header_index, "差旅-住宿（万元）", "差旅-住宿")
        travel_allowance = number_by_header(row, header_index, "差旅-差补（万元）", "差旅-差补")
        travel_total = number_by_header(row, header_index, "差旅合计（万元）", "差旅合计")
        if not travel_total:
            travel_total = travel_transportation + travel_lodging + travel_allowance
        records.append({
            "person_name": person_name,
            "department": department,
            "mdm_project_name": mdm_project_name,
            "travel_transportation_amount": travel_transportation,
            "travel_lodging_amount": travel_lodging,
            "travel_allowance_amount": travel_allowance,
            "travel_total_amount": travel_total,
        })
    return records


def parse_person_hospitality_rows(wb: Any) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    ws = resolve_sheet(wb, PERSON_HOSPITALITY_SHEET)
    header_row, header_index = find_header_row(ws, ["人员", "MDM项目名称", "招待性质"])
    for row in rows(ws, header_row + 1):
        person_name = clean_text(value_by_header(row, header_index, "人员", "姓名"))
        department = clean_text(value_by_header(row, header_index, "当前所属部门", "所属部门"))
        mdm_project_name = clean_text(value_by_header(row, header_index, "MDM项目名称"))
        hospitality_type = clean_text(value_by_header(row, header_index, "招待性质"))
        if not person_name or not mdm_project_name or not hospitality_type:
            continue
        guest_count = number_by_header(row, header_index, "接待人数")
        amount = number_by_header(row, header_index, "业务招待费金额（万元）", "业务招待费（万元）", "业务招待合计（万元）")
        per_capita_amount = to_number(value_by_header(row, header_index, "人均标准"))
        records.append({
            "person_name": person_name,
            "department": department,
            "mdm_project_name": mdm_project_name,
            "hospitality_type": hospitality_type,
            "guest_count": guest_count,
            "hospitality_total_amount": amount,
            "per_capita_amount": per_capita_amount if per_capita_amount is not None else (amount / guest_count if guest_count else None),
        })
    return records


def parse_project_summary_rows(wb: Any) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    ws = resolve_sheet(wb, PROJECT_SUMMARY_SHEET)
    header_row, header_index = find_header_row(ws, ["项目标签", "所属区域"])
    for row in rows(ws, header_row + 1):
        project_tag = clean_text(value_by_header(row, header_index, "项目标签"))
        if not project_tag or "合计" in project_tag:
            continue
        travel_transportation = number_by_header(row, header_index, "差旅-交通（万元）", "差旅-交通")
        travel_lodging = number_by_header(row, header_index, "差旅-住宿（万元）", "差旅-住宿")
        travel_allowance = number_by_header(row, header_index, "差旅-差补（万元）", "差旅-差补")
        travel_total = number_by_header(row, header_index, "差旅合计（万元）", "差旅合计")
        if not travel_total:
            travel_total = travel_transportation + travel_lodging + travel_allowance
        hospitality_amount = number_by_header(row, header_index, "业务招待费（万元）", "业务招待合计（万元）")
        bonus_amount = number_by_header(row, header_index, "已发市场奖金（万元）", "已发市场奖金")
        total_expense = number_by_header(row, header_index, "费用合计", "差旅招待总计（万元）")
        if not total_expense:
            total_expense = travel_total + hospitality_amount + bonus_amount
        profit = number_by_header(row, header_index, "立项首年利润额（万元）", "立项首年利润额")
        roi = to_number(value_by_header(row, header_index, "首年ROI\n（产出/投入）", "首年ROI（产出/投入）", "首年ROI"))
        records.append({
            "project_tag": project_tag,
            "region": clean_text(value_by_header(row, header_index, "所属区域")),
            "launch_date": to_date(value_by_header(row, header_index, "立项时间")),
            "first_year_contract_amount": number_by_header(row, header_index, "立项首年合同额（万元）", "立项首年合同额"),
            "first_year_profit_amount": profit,
            "travel_transportation_amount": travel_transportation,
            "travel_lodging_amount": travel_lodging,
            "travel_allowance_amount": travel_allowance,
            "travel_total_amount": travel_total,
            "hospitality_total_amount": hospitality_amount,
            "paid_market_bonus_amount": bonus_amount,
            "total_expense_amount": total_expense,
            "first_year_roi": roi if roi is not None else (profit / total_expense if total_expense else None),
        })
    return records


def attach_batch_id(records: list[dict[str, Any]], batch_id: str) -> list[dict[str, Any]]:
    return [{**record, "batch_id": batch_id} for record in records]


def request_with_retries(send_request: Callable[[], httpx.Response], action_label: str) -> httpx.Response:
    for attempt in range(1, REQUEST_RETRIES + 1):
        try:
            return send_request()
        except (httpx.ConnectTimeout, httpx.ReadTimeout, httpx.WriteTimeout, httpx.ConnectError) as error:
            if attempt >= REQUEST_RETRIES:
                print(f"{action_label} 失败: {type(error).__name__}: {error}")
                sys.exit(1)
            wait_seconds = attempt * 3
            print(f"{action_label} 超时/连接失败，{wait_seconds} 秒后重试 ({attempt}/{REQUEST_RETRIES})...")
            time.sleep(wait_seconds)
    raise RuntimeError("unreachable")


def explain_http_error(error: httpx.HTTPStatusError, table: str) -> None:
    response = error.response
    print(f"{table} 写入失败: HTTP {response.status_code}")
    print(response.text)
    if response.status_code == 404:
        print("提示: 目标表不存在，请先在 Supabase 应用最新 migrations。")
    elif response.status_code in (401, 403):
        print("提示: 当前 key 没有写入权限。若费效表未关闭 RLS，请先应用关闭 RLS 的迁移，或设置 SUPABASE_SERVICE_ROLE_KEY。")


def post_json(client: httpx.Client, supabase_url: str, table: str, records: list[dict[str, Any]], return_representation: bool = False) -> list[dict[str, Any]]:
    if not records:
        return []
    inserted: list[dict[str, Any]] = []
    headers = {"Prefer": "return=representation"} if return_representation else {"Prefer": "return=minimal"}
    for index in range(0, len(records), BATCH_SIZE):
        batch = records[index : index + BATCH_SIZE]
        response = request_with_retries(
            lambda: client.post(f"{supabase_url}/rest/v1/{table}", json=batch, headers=headers),
            f"{table} 批次 {index // BATCH_SIZE + 1} 写入",
        )
        try:
            response.raise_for_status()
        except httpx.HTTPStatusError as error:
            explain_http_error(error, table)
            sys.exit(1)
        if return_representation:
            inserted.extend(response.json())
        print(f"{table}: 批次 {index // BATCH_SIZE + 1} 写入 {len(batch)} 条")
    return inserted


def delete_existing_batches(client: httpx.Client, supabase_url: str) -> None:
    response = request_with_retries(
        lambda: client.delete(f"{supabase_url}/rest/v1/{BATCH_TABLE}", params={"id": "not.is.null"}),
        "清空旧费效批次",
    )
    if response.status_code not in (200, 204):
        print(f"清空旧费效批次失败: {response.status_code} {response.text}")
        sys.exit(1)


def main() -> None:
    args = parse_args()
    excel_path = Path(args.input).resolve()
    if not excel_path.exists():
        print(f"Excel 文件不存在: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    period_start, period_end = period_range(wb)
    person_summary_rows = parse_person_summary_rows(wb)
    person_travel_rows = parse_person_travel_rows(wb)
    person_hospitality_rows = parse_person_hospitality_rows(wb)
    project_summary_rows = parse_project_summary_rows(wb)

    print(f"读取文件: {excel_path}")
    print(f"统计期间: {period_start}-{period_end}")
    print(f"{PERSON_SUMMARY_SHEET}: {len(person_summary_rows)} 行")
    print(f"{PERSON_TRAVEL_SHEET}: {len(person_travel_rows)} 行")
    print(f"{PERSON_HOSPITALITY_SHEET}: {len(person_hospitality_rows)} 行")
    print(f"{PROJECT_SUMMARY_SHEET}: {len(project_summary_rows)} 行")

    if args.dry_run:
        print("Dry run 完成，未写入数据库。")
        return
    if not args.confirm:
        print("检测到将执行写入操作。请追加 --confirm 后重试。")
        sys.exit(2)

    supabase_url, write_key, auth_mode = require_supabase_env()
    headers = {
        "apikey": write_key,
        "Authorization": f"Bearer {write_key}",
        "Content-Type": "application/json",
    }
    print(f"Supabase 写入认证模式: {auth_mode}")
    if auth_mode == "anon":
        print("提示: 当前使用 VITE_SUPABASE_ANON_KEY 写入，要求费效表已关闭 RLS 且具备 REST 写入权限。")

    timeout = httpx.Timeout(connect=45, read=120, write=120, pool=45)
    with httpx.Client(headers=headers, timeout=timeout) as client:
        if not args.no_clear:
            print("正在清空旧费效批次...")
            delete_existing_batches(client, supabase_url)

        batch_record = {
            "source_file_name": excel_path.name,
            "source_file_hash": file_hash(excel_path),
            "period_start": period_start,
            "period_end": period_end,
            "person_summary_row_count": len(person_summary_rows),
            "person_travel_row_count": len(person_travel_rows),
            "person_hospitality_row_count": len(person_hospitality_rows),
            "project_summary_row_count": len(project_summary_rows),
        }
        inserted_batch = post_json(client, supabase_url, BATCH_TABLE, [batch_record], return_representation=True)
        batch_id = inserted_batch[0]["id"]

        post_json(client, supabase_url, "fee_effect_person_summary", attach_batch_id(person_summary_rows, batch_id))
        post_json(client, supabase_url, "fee_effect_person_travel_projects", attach_batch_id(person_travel_rows, batch_id))
        post_json(client, supabase_url, "fee_effect_person_hospitality_projects", attach_batch_id(person_hospitality_rows, batch_id))
        post_json(client, supabase_url, "fee_effect_project_summary", attach_batch_id(project_summary_rows, batch_id))

    print(f"费效数据导入完成，batch_id={batch_id}")


if __name__ == "__main__":
    main()
