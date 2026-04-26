"""
读取 25学年经营数据.xlsx 中的多个 sheet 页，解析经营数据并写入 Supabase 数据库。

包含的 sheet 页：
- 1.x: fone 版经营数据（累计/月度）
- 2.x: 突围版经营数据（累计/月度）
- 3.x: 突围版成本分析（累计/月度）
- 4.x: fone 版成本分析（累计/月度）
- 5: 5年战略预算规划

同时读取组织标签映射表，创建独立的 edu_org_hierarchy 表。
数据表之间无外键约束，仅通过 node_name 字段关联。

注意：人力成本指标在主报表和成本分析表中都存在，导入时会自动去重，
优先保留成本分析表（3.x/4.x）中的数据。

用法:
  python scripts/import_biz_data.py --dry-run
  python scripts/import_biz_data.py --confirm
"""

import argparse
import os
import sys
import json
import re
import openpyxl
import requests
from pathlib import Path

# ─── 配置 ───────────────────────────────────────────────
ROOT_DIR = Path(__file__).resolve().parent.parent
DEFAULT_EXCEL_PATH = ROOT_DIR / "private-data" / "25学年经营数据.xlsx"
DEFAULT_ORG_HIERARCHY_PATH = ROOT_DIR / "private-data" / "教育后勤2025经营数据看版_组织标签映射表-勿动.xlsx"


def parse_args():
    parser = argparse.ArgumentParser(description="导入经营数据与组织映射到 Supabase")
    parser.add_argument("--input", default=str(DEFAULT_EXCEL_PATH), help="经营数据 Excel 文件路径")
    parser.add_argument("--org-map", default=str(DEFAULT_ORG_HIERARCHY_PATH), help="组织映射 Excel 文件路径")
    parser.add_argument("--dry-run", action="store_true", help="仅解析并输出摘要，不写入数据库")
    parser.add_argument("--confirm", action="store_true", help="确认执行覆盖写入")
    return parser.parse_args()


def load_env():
    env_path = Path(__file__).parent.parent / "app" / ".env"
    env = {}
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip().strip('"').strip("'")
    return env

def require_supabase_env():
    env = load_env()
    supabase_url = os.getenv("SUPABASE_URL") or env.get("SUPABASE_URL") or env.get("VITE_SUPABASE_URL", "")
    service_role_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or env.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not supabase_url or not service_role_key:
        print("错误: 未找到 SUPABASE_URL 或 SUPABASE_SERVICE_ROLE_KEY，请检查 app/.env 或当前环境变量")
        sys.exit(1)
    return {
        "SUPABASE_URL": supabase_url,
        "HEADERS": {
            "apikey": service_role_key,
            "Authorization": f"Bearer {service_role_key}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        },
    }

# ─── Sheet 类型映射（按主编号）────────────────────────────
MAIN_SHEET_REPORT_TYPES = {
    "1": "fone",
    "2": "tuwei",
}

COST_SHEET_REPORT_TYPES = {
    "3": "tuwei",
    "4": "fone",
}

# 16 个指标大类，每类占 5 列（从 C2 开始）
METRIC_CATEGORIES = [
    ("revenue",           "营业收入",    2),
    ("catering_expense",  "餐饮支出",    7),
    ("material_cost",     "物资销售成本", 12),
    ("gross_profit",      "毛利额",      17),
    ("gross_margin",      "毛利率",      22),
    ("labor_cost",        "人力成本",    27),
    ("other_expense",     "其他支出",    32),
    ("external_revenue",  "营业外收入",  37),
    ("external_expense",  "营业外支出",  42),
    ("pretax_profit",     "税前利润",    47),
    ("pretax_margin",     "税前利润率",  52),
    ("headcount",         "职工人数",    57),
    ("per_capita_revenue","人均营收",    62),
    ("labor_cost_rate",   "人力成本率",  67),
    ("revenue_creation",  "一元创收",    72),
    ("profit_creation",   "一元创利",    77),
]

# Sheet 2.1 末尾附加列：突围计划 1-6 整体目标 / 完成进度
# 结合该 sheet 原有的累计实际值，生成 period=<202607 的 cumulative 指标
TUWEI_CUMULATIVE_EXTRA_METRICS = [
    {
        "metric_category": "revenue",
        "metric_category_cn": "营业收入",
        "actual_col": 2,
        "budget_col": 72,
        "completion_col": 74,
    },
    {
        "metric_category": "pretax_profit",
        "metric_category_cn": "税前利润",
        "actual_col": 47,
        "budget_col": 73,
        "completion_col": 75,
    },
]

# 成本分析指标（6.1/6.2/7.1/7.2），每类占 5 列（实际值、考核数、预算完成率、预实差异、同期）
COST_METRIC_CATEGORIES = [
    ("labor_cost",        "人力成本",      2),
    ("salary",            "工资",          7),
    ("social_insurance",  "社保",          12),
    ("housing_fund",      "公积金",        17),
    ("labor_service_fee", "劳务费",        22),
    ("other_labor_cost",  "其他人力成本",  27),
    ("vehicle_expense",   "车辆费用",      32),
    ("energy_expense",    "能耗费",        37),
    ("travel_expense",    "差旅费",        42),
    ("entertainment_expense", "业务招待费", 47),
]

DATA_ROW_START = 8
DATA_ROW_END = 139

STRATEGY_PLAN_TABLE = "edu_strategy_budget_plan"
STRATEGY_PLAN_SHEET_CODE = "5"
STRATEGY_YEAR_COLUMNS = [
    (2025, 3, 4),
    (2026, 5, 6),
    (2027, 7, 8),
    (2028, 9, 10),
    (2029, 11, 12),
    (2030, 13, 14),
]


def normalize_header(value) -> str:
    """Normalize Excel header text for header-based column lookup."""
    return str(value or "").strip().lower()


def load_org_hierarchy(excel_path: Path) -> list[dict]:
    """
    从组织标签映射表加载组织层级数据
    返回: [{node_name, level_0, level_1, level_2}, ...]
    """
    if not excel_path.exists():
        print(f"  警告: 组织标签映射文件不存在 {excel_path}")
        return []

    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    ws = wb.active

    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(row=1, column=col_idx).value)
        if header:
            header_map[header] = col_idx

    required_headers = ["level_0", "level_1", "level_2", "node_name"]
    missing_headers = [header for header in required_headers if header not in header_map]
    if missing_headers:
        raise ValueError(f"组织标签映射表缺少字段: {', '.join(missing_headers)}")

    rows_by_node = {}
    for row_idx in range(2, ws.max_row + 1):
        node_name = str(ws.cell(row=row_idx, column=header_map["node_name"]).value or "").strip()
        if not node_name:
            continue

        level_0 = str(ws.cell(row=row_idx, column=header_map["level_0"]).value or "").strip() or None
        level_1 = str(ws.cell(row=row_idx, column=header_map["level_1"]).value or "").strip() or None
        level_2 = str(ws.cell(row=row_idx, column=header_map["level_2"]).value or "").strip() or None

        rows_by_node[node_name] = {
            "node_name": node_name,
            "level_0": level_0,
            "level_1": level_1,
            "level_2": level_2,
        }

    wb.close()
    rows = list(rows_by_node.values())
    print(f"  加载了 {len(rows)} 个组织节点")
    return rows


def clean_org_hierarchy(rows: list[dict]) -> list[dict]:
    """
    清理组织层级数据：
    1. 如果不同层级出现相同名称，将低层级的值设为空
    2. 如果 level_0-2 的值与 node_name 相同，将对应的 level 值设为空
    """
    cleaned_count = 0

    for row in rows:
        node_name = row["node_name"]
        level_0 = row["level_0"]
        level_1 = row["level_1"]
        level_2 = row["level_2"]

        if level_0 and level_0 == node_name:
            row["level_0"] = None
            cleaned_count += 1

        if level_1 and level_1 == node_name:
            row["level_1"] = None
            cleaned_count += 1

        if level_2 and level_2 == node_name:
            row["level_2"] = None
            cleaned_count += 1

        if level_0 and row["level_1"] and level_0 == row["level_1"]:
            row["level_1"] = None
            cleaned_count += 1

        if row["level_1"] and row["level_2"] and row["level_1"] == row["level_2"]:
            row["level_2"] = None
            cleaned_count += 1

    print(f"  清理了 {cleaned_count} 个重复层级值")
    return rows


def safe_num(v):
    """将 Excel 单元格值转为 float 或 None"""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v in ("", "/", "-", "--", "—"):
            return None
        try:
            return float(v.replace(",", ""))
        except ValueError:
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def safe_text(v) -> str | None:
    """将 Excel 单元格值转为去空白后的文本或 None"""
    if v is None:
        return None
    text = str(v).strip()
    return text or None


def detect_period_type(raw_period) -> str | None:
    """
    根据期间单元格内容识别期间类型。
    - <202604: cumulative
    - 202603: monthly
    - 202601-202603 / 202601~202603: cumulative
    """
    text = str(raw_period or "").strip()
    if not text:
        return None

    if re.search(r"<\s*20\d{4}", text):
        return "cumulative"

    months = re.findall(r"20\d{4}", text)
    if len(months) >= 2:
        return "cumulative"
    if len(months) == 1:
        return "monthly"

    return None


def detect_sheet_period_type(ws) -> str:
    """
    优先根据期间列自动识别 sheet 的期间类型。
    若 B3/C3 都无法识别，则回退为 monthly，避免整张表中断。
    """
    for col in (2, 3):
        detected = detect_period_type(ws.cell(row=3, column=col).value)
        if detected:
            return detected

    return "monthly"


def normalize_period(raw_period, period_type: str | None = None) -> str | None:
    """
    统一期间字段格式。
    - monthly: 保持 YYYYMM
    - cumulative: 统一为 <下月YYYYMM（右开区间）>
    """
    text = str(raw_period or "").strip()
    if not text:
        return None

    period_type = period_type or detect_period_type(text)
    if not period_type:
        return text

    exclusive_match = re.search(r"<\s*(20\d{4})", text)
    if period_type == "cumulative" and exclusive_match:
        return f"<{exclusive_match.group(1)}"

    months = re.findall(r"20\d{4}", text)
    if not months:
        return text

    if period_type == "cumulative":
        last_month = months[-1]
        year = int(last_month[:4])
        month = int(last_month[4:])
        if not 1 <= month <= 12:
            return f"<{last_month}"

        month += 1
        if month == 13:
            year += 1
            month = 1

        return f"<{year}{month:02d}"

    if len(months) == 1:
        return months[0]

    return text


def period_sort_key(period: str | None) -> int:
    """将期间文本转成可比较的排序值，越大表示期间越新。"""
    if not period:
        return -1

    match = re.search(r"20\d{4}", period)
    if not match:
        return -1

    return int(match.group(0))


def extract_sheet_code(sheet_name: str) -> str | None:
    """从 sheet 名开头提取编号，如 1.4、2.5、5。"""
    match = re.match(r"^\s*(\d+(?:\.\d+)?)", str(sheet_name or "").strip())
    if not match:
        return None
    return match.group(1)


def sheet_code_sort_key(sheet_code: str) -> tuple[int, int]:
    """按编号排序，支持 1.4、2.5、5。"""
    parts = str(sheet_code).split(".", 1)
    major = int(parts[0])
    minor = int(parts[1]) if len(parts) > 1 else 0
    return (major, minor)


def select_sheet_name(wb, sheet_code: str) -> str | None:
    """
    选择与 sheet_code 匹配的 sheet。
    - 若只有一个匹配项，直接返回
    - 若有多个匹配项，优先按 Row 3 / Col 2 的期间信息选择最新版本
    - 若无法比较期间，则回退到最后一个匹配项
    """
    matched = [s for s in wb.sheetnames if s.startswith(sheet_code)]
    if not matched:
        return None
    if len(matched) == 1:
        return matched[0]

    ranked = []
    for sheet_name in matched:
        ws = wb[sheet_name]
        period_type = detect_sheet_period_type(ws)
        period = normalize_period(ws.cell(row=3, column=2).value, period_type)
        ranked.append((period_sort_key(period), sheet_name))

    ranked.sort(key=lambda item: item[0], reverse=True)
    best_key, best_sheet = ranked[0]
    if best_key >= 0:
        print(f"  检测到多个 sheet 匹配 {sheet_code}，使用期间最新的 [{best_sheet}]")
        return best_sheet

    best_sheet = matched[-1]
    print(f"  检测到多个 sheet 匹配 {sheet_code}，回退使用最后一个 [{best_sheet}]")
    return best_sheet


def discover_sheet_codes(wb, allowed_major_codes: set[str]) -> list[str]:
    """
    扫描 workbook 中所有 sheet，提取允许的编号，并去重排序。
    如果同一编号存在多个版本，仅返回一次，后续由 select_sheet_name 选最新版本。
    """
    discovered = set()
    for sheet_name in wb.sheetnames:
        sheet_code = extract_sheet_code(sheet_name)
        if not sheet_code:
            continue
        major_code = sheet_code.split(".", 1)[0]
        if major_code in allowed_major_codes:
            discovered.add(sheet_code)

    return sorted(discovered, key=sheet_code_sort_key)


def upsert_batch(supabase_url: str, headers: dict, table: str, rows: list[dict], batch_size: int = 500):
    """分批写入 Supabase"""
    url = f"{supabase_url}/rest/v1/{table}"
    total = 0
    for i in range(0, len(rows), batch_size):
        batch = rows[i : i + batch_size]
        resp = requests.post(url, headers=headers, data=json.dumps(batch))
        if resp.status_code not in (200, 201):
            print(f"  写入失败 (batch {i // batch_size + 1}): {resp.status_code}")
            print(f"  响应: {resp.text[:500]}")
            return total
        total += len(batch)
    return total


def clear_table(supabase_url: str, headers: dict, table: str):
    """清空目标表"""
    url = f"{supabase_url}/rest/v1/{table}?id=not.is.null"
    resp = requests.delete(url, headers=headers)
    if resp.status_code in (200, 204):
        print(f"  已清空 {table}")
    else:
        print(f"  清空 {table} 失败: {resp.status_code} {resp.text[:200]}")


def infer_strategy_group(line_label: str) -> tuple[str, str]:
    """为战略预算行识别分组，便于咨询分析与前端过滤。"""
    if line_label.startswith("基本盘"):
        return ("base_business", "基本盘")
    if line_label.startswith("增长极"):
        return ("growth_engine", "增长极")
    if line_label == "合计":
        return ("overall_total", "合计")
    return ("strategic_kpi", "战略指标")


def infer_strategy_line_role(line_label: str) -> str:
    """识别行角色：明细/小计/合计/战略指标。"""
    if line_label == "合计":
        return "total"
    if line_label.endswith("小计"):
        return "subtotal"
    if any(keyword in line_label for keyword in ("增速", "利润率", "成本率")):
        return "kpi"
    return "detail"


def normalize_strategy_business_line(line_label: str) -> str:
    """从指标标签中提取标准业务条线名称。"""
    mapping = [
        ("三大区域营收增速", "三大区域"),
        ("三大区域利润率", "三大区域"),
        ("集团本级管理部门成本率", "集团本级管理部门"),
        ("商超业务利润率", "商超业务"),
    ]
    for source, normalized in mapping:
        if line_label == source:
            return normalized
    return line_label


def parse_strategy_kpi_definition(line_label: str) -> tuple[str, str, str, str] | None:
    """
    从战略 KPI 行名识别指标定义：
    返回 (metric_code, metric_name_cn, unit, value_type)
    """
    definitions = {
        "三大区域营收增速": ("revenue_growth_rate", "营收增速", "ratio", "ratio"),
        "三大区域利润率": ("profit_margin", "利润率", "ratio", "ratio"),
        "集团本级管理部门成本率": ("cost_ratio", "成本率", "ratio", "ratio"),
        "商超业务利润率": ("profit_margin", "利润率", "ratio", "ratio"),
    }
    return definitions.get(line_label)


def parse_strategy_budget_sheet(wb) -> list[dict]:
    """解析 sheet 5：5年战略预算规划，按 条线 × 年份 × 指标 落成长表。"""
    sheet_name = select_sheet_name(wb, STRATEGY_PLAN_SHEET_CODE)
    if not sheet_name:
        print("  警告: 未找到 sheet 5，跳过战略预算规划导入")
        return []

    ws = wb[sheet_name]
    print(f"  处理 [{sheet_name}]")

    source_note = safe_text(ws.cell(row=14, column=2).value)
    all_rows = []

    for row_idx in range(2, ws.max_row + 1):
        line_label = safe_text(ws.cell(row=row_idx, column=2).value)
        if not line_label or line_label.startswith("注明"):
            continue

        strategy_group, strategy_group_cn = infer_strategy_group(line_label)
        line_role = infer_strategy_line_role(line_label)
        business_line = normalize_strategy_business_line(line_label)

        if line_role == "kpi":
            kpi_definition = parse_strategy_kpi_definition(line_label)
            if not kpi_definition:
                print(f"    警告: 未识别的战略 KPI 行 [{line_label}]，已跳过")
                continue

            metric_code, metric_name_cn, unit, value_type = kpi_definition
            for col_idx in range(3, ws.max_column + 1):
                value = safe_num(ws.cell(row=row_idx, column=col_idx).value)
                if value is None:
                    continue

                header = safe_text(ws.cell(row=1, column=col_idx).value)
                year_match = re.search(r"(20\d{2})自然年", header or "")
                if not year_match:
                    continue

                all_rows.append(
                    {
                        "strategy_group": strategy_group,
                        "strategy_group_cn": strategy_group_cn,
                        "line_role": line_role,
                        "business_line": business_line,
                        "line_label": line_label,
                        "plan_year": int(year_match.group(1)),
                        "metric_code": metric_code,
                        "metric_name_cn": metric_name_cn,
                        "value": value,
                        "unit": unit,
                        "value_type": value_type,
                        "sort_order": row_idx,
                        "source_note": source_note,
                    }
                )
            continue

        for plan_year, revenue_col, profit_col in STRATEGY_YEAR_COLUMNS:
            revenue = safe_num(ws.cell(row=row_idx, column=revenue_col).value)
            if revenue is not None:
                all_rows.append(
                    {
                        "strategy_group": strategy_group,
                        "strategy_group_cn": strategy_group_cn,
                        "line_role": line_role,
                        "business_line": business_line,
                        "line_label": line_label,
                        "plan_year": plan_year,
                        "metric_code": "revenue",
                        "metric_name_cn": "营收",
                        "value": revenue,
                        "unit": "amount",
                        "value_type": "absolute",
                        "sort_order": row_idx,
                        "source_note": source_note,
                    }
                )

            profit = safe_num(ws.cell(row=row_idx, column=profit_col).value)
            if profit is not None:
                all_rows.append(
                    {
                        "strategy_group": strategy_group,
                        "strategy_group_cn": strategy_group_cn,
                        "line_role": line_role,
                        "business_line": business_line,
                        "line_label": line_label,
                        "plan_year": plan_year,
                        "metric_code": "profit",
                        "metric_name_cn": "利润",
                        "value": profit,
                        "unit": "amount",
                        "value_type": "absolute",
                        "sort_order": row_idx,
                        "source_note": source_note,
                    }
                )

    print(f"    -> {len(all_rows)} 条战略预算规划数据")
    return all_rows


def parse_main_sheets(wb, valid_nodes: set):
    """解析经营数据 sheets 1.x-2.x，只导入映射表中存在的节点"""
    all_rows = []
    skipped_nodes = set()
    sheet_codes = discover_sheet_codes(wb, set(MAIN_SHEET_REPORT_TYPES))
    if not sheet_codes:
        print("  警告: 未找到经营数据 sheet（1.x/2.x）")
        return []

    for sheet_code in sheet_codes:
        major_code = sheet_code.split(".", 1)[0]
        report_type = MAIN_SHEET_REPORT_TYPES.get(major_code)
        if not report_type:
            continue

        sheet_name = select_sheet_name(wb, sheet_code)
        if not sheet_name:
            print(f"  警告: 未找到 sheet {sheet_code}")
            continue
        ws = wb[sheet_name]
        print(f"  处理 [{sheet_name}]")

        # 读取期间信息 (Row 3)
        period_type = detect_sheet_period_type(ws)
        period = normalize_period(ws.cell(row=3, column=2).value, period_type)
        period_yoy = normalize_period(ws.cell(row=3, column=3).value, period_type)

        sheet_rows = 0
        for row_idx in range(DATA_ROW_START, DATA_ROW_END + 1):
            node_name = ws.cell(row=row_idx, column=1).value
            if node_name is None:
                continue
            node_name = str(node_name).strip()
            if not node_name:
                continue

            # 只导入映射表中存在的节点
            if node_name not in valid_nodes:
                skipped_nodes.add(node_name)
                continue

            for metric_en, metric_cn, start_col in METRIC_CATEGORIES:
                actual = safe_num(ws.cell(row=row_idx, column=start_col).value)
                budget = safe_num(ws.cell(row=row_idx, column=start_col + 1).value)
                rate = safe_num(ws.cell(row=row_idx, column=start_col + 2).value)
                diff = safe_num(ws.cell(row=row_idx, column=start_col + 3).value)

                # 人均营收只有4列（无同期列），其余有5列
                if metric_en == "per_capita_revenue":
                    yoy = None
                elif metric_en == "profit_creation":
                    yoy = None  # 一元创利也只有4列
                else:
                    yoy = safe_num(ws.cell(row=row_idx, column=start_col + 4).value)

                # 跳过全部为空的指标行
                if all(v is None for v in [actual, budget, rate, diff, yoy]):
                    continue

                row_data = {
                    "sheet_code": sheet_code,
                    "report_type": report_type,
                    "period_type": period_type,
                    "period": period,
                    "period_yoy": period_yoy,
                    "node_name": node_name,
                    "metric_category": metric_en,
                    "metric_category_cn": metric_cn,
                    "actual_value": actual,
                    "budget_value": budget,
                    "completion_rate": rate,
                    "diff_value": diff,
                    "yoy_value": yoy,
                    "sort_order": row_idx,
                }

                all_rows.append(row_data)
                sheet_rows += 1

        print(f"    -> {sheet_rows} 条指标数据")

    if skipped_nodes:
        print(f"\n  跳过了 {len(skipped_nodes)} 个不在映射表中的节点:")
        for node in sorted(skipped_nodes):
            print(f"    - {node}")

    return all_rows


def parse_tuwei_cumulative_extra_metrics(wb, valid_nodes: set):
    """解析 2.1 突围累计 sheet 末尾整体目标列，生成 <202607 的 cumulative 指标。"""
    sheet_name = select_sheet_name(wb, "2.1")
    if not sheet_name:
        print("  警告: 未找到 sheet 2.1，跳过突围累计附加指标")
        return []

    ws = wb[sheet_name]
    print(f"  处理 [{sheet_name}] 的突围累计附加指标")

    all_rows = []
    skipped_nodes = set()

    for row_idx in range(DATA_ROW_START, DATA_ROW_END + 1):
        node_name = ws.cell(row=row_idx, column=1).value
        if node_name is None:
            continue
        node_name = str(node_name).strip()
        if not node_name:
            continue

        if node_name not in valid_nodes:
            skipped_nodes.add(node_name)
            continue

        for metric in TUWEI_CUMULATIVE_EXTRA_METRICS:
            actual = safe_num(ws.cell(row=row_idx, column=metric["actual_col"]).value)
            budget = safe_num(ws.cell(row=row_idx, column=metric["budget_col"]).value)
            completion = safe_num(ws.cell(row=row_idx, column=metric["completion_col"]).value)

            if all(v is None for v in [actual, budget, completion]):
                continue

            all_rows.append(
                {
                    "sheet_code": "2.1",
                    "report_type": "tuwei",
                    "period_type": "cumulative",
                    "period": "<202607",
                    "period_yoy": None,
                    "node_name": node_name,
                    "metric_category": metric["metric_category"],
                    "metric_category_cn": metric["metric_category_cn"],
                    "actual_value": actual,
                    "budget_value": budget,
                    "completion_rate": completion,
                    "diff_value": actual - budget if actual is not None and budget is not None else None,
                    "yoy_value": None,
                    "sort_order": row_idx,
                }
            )

    print(f"    -> {len(all_rows)} 条突围累计附加指标数据")

    if skipped_nodes:
        print(f"\n  跳过了 {len(skipped_nodes)} 个不在映射表中的节点:")
        for node in sorted(skipped_nodes):
            print(f"    - {node}")

    return all_rows


def parse_cost_sheets(wb, valid_nodes: set):
    """解析成本分析 sheets 3.x-4.x，只导入映射表中存在的节点"""
    all_rows = []
    skipped_nodes = set()
    sheet_codes = discover_sheet_codes(wb, set(COST_SHEET_REPORT_TYPES))
    if not sheet_codes:
        print("  警告: 未找到成本分析 sheet（3.x/4.x）")
        return []

    for sheet_code in sheet_codes:
        major_code = sheet_code.split(".", 1)[0]
        report_type = COST_SHEET_REPORT_TYPES.get(major_code)
        if not report_type:
            continue

        sheet_name = select_sheet_name(wb, sheet_code)
        if not sheet_name:
            print(f"  警告: 未找到 sheet {sheet_code}")
            continue
        ws = wb[sheet_name]
        print(f"  处理 [{sheet_name}]")

        # 读取期间信息 (Row 3)
        period_type = detect_sheet_period_type(ws)
        period = normalize_period(ws.cell(row=3, column=2).value, period_type)
        period_yoy = normalize_period(ws.cell(row=3, column=3).value, period_type)

        sheet_rows = 0
        for row_idx in range(DATA_ROW_START, DATA_ROW_END + 1):
            node_name = ws.cell(row=row_idx, column=1).value
            if node_name is None:
                continue
            node_name = str(node_name).strip()
            if not node_name:
                continue

            # 只导入映射表中存在的节点
            if node_name not in valid_nodes:
                skipped_nodes.add(node_name)
                continue

            for metric_en, metric_cn, start_col in COST_METRIC_CATEGORIES:
                actual = safe_num(ws.cell(row=row_idx, column=start_col).value)
                budget = safe_num(ws.cell(row=row_idx, column=start_col + 1).value)
                rate = safe_num(ws.cell(row=row_idx, column=start_col + 2).value)
                diff = safe_num(ws.cell(row=row_idx, column=start_col + 3).value)
                yoy = safe_num(ws.cell(row=row_idx, column=start_col + 4).value)

                # 跳过全部为空的指标行
                if all(v is None for v in [actual, budget, rate, diff, yoy]):
                    continue

                row_data = {
                    "sheet_code": sheet_code,
                    "report_type": report_type,
                    "period_type": period_type,
                    "period": period,
                    "period_yoy": period_yoy,
                    "node_name": node_name,
                    "metric_category": metric_en,
                    "metric_category_cn": metric_cn,
                    "actual_value": actual,
                    "budget_value": budget,
                    "completion_rate": rate,
                    "diff_value": diff,
                    "yoy_value": yoy,
                    "sort_order": row_idx,
                }

                all_rows.append(row_data)
                sheet_rows += 1

        print(f"    -> {sheet_rows} 条成本指标数据")

    if skipped_nodes:
        print(f"\n  跳过了 {len(skipped_nodes)} 个不在映射表中的节点:")
        for node in sorted(skipped_nodes):
            print(f"    - {node}")

    return all_rows


def main():
    args = parse_args()
    excel_path = Path(args.input).resolve()
    org_hierarchy_path = Path(args.org_map).resolve()
    print(f"读取 Excel: {excel_path}")
    if not excel_path.exists():
        print(f"错误: 文件不存在 {excel_path}")
        sys.exit(1)
    if not org_hierarchy_path.exists():
        print(f"错误: 组织映射文件不存在 {org_hierarchy_path}")
        sys.exit(1)

    # 加载组织层级数据
    print(f"\n加载组织层级映射: {org_hierarchy_path}")
    org_hierarchy_rows = load_org_hierarchy(org_hierarchy_path)

    # 清理组织层级数据
    print("\n清理组织层级数据...")
    org_hierarchy_rows = clean_org_hierarchy(org_hierarchy_rows)

    # 构建有效节点集合
    valid_nodes = {row["node_name"] for row in org_hierarchy_rows}
    print(f"  映射表中有 {len(valid_nodes)} 个有效节点")

    wb = openpyxl.load_workbook(str(excel_path), data_only=True, keep_links=False)
    print(f"共 {len(wb.sheetnames)} 个 sheet\n")

    # 解析主报表（只导入映射表中存在的节点）
    print("解析经营数据报表 (1.x-2.x)...")
    report_rows = parse_main_sheets(wb, valid_nodes)
    print(f"\n共 {len(report_rows)} 条报表数据")

    # 解析成本分析报表 (3.x-4.x)
    print("\n解析成本分析报表 (3.x-4.x)...")
    cost_rows = parse_cost_sheets(wb, valid_nodes)
    print(f"共 {len(cost_rows)} 条成本数据")

    # 合并报表数据和成本数据，去重人力成本
    print("\n合并数据并去重人力成本...")
    all_report_rows = report_rows + cost_rows

    # 去重逻辑：对于相同的 (report_type, period_type, node_name, metric_category)
    # 如果 metric_category 是 labor_cost，优先保留成本分析表中的数据（来自 3.x/4.x）
    # 因为成本分析表提供了更详细的人力成本分解
    dedup_key_map = {}
    for row in all_report_rows:
        # 对于人力成本，使用不含 sheet_code 的 key 进行去重
        if row["metric_category"] == "labor_cost":
            key = (
                row["report_type"],
                row["period_type"],
                row["node_name"],
                row["metric_category"],
            )

            if key in dedup_key_map:
                current_sheet = row["sheet_code"]
                # 如果当前是成本分析表（3.x/4.x），替换之前的数据
                if current_sheet.startswith(("3", "4")):
                    dedup_key_map[key] = row
                # 如果已存在的是成本分析表，保持不变（不添加当前行）
            else:
                dedup_key_map[key] = row
        else:
            # 非人力成本指标，使用完整 key（包含 sheet_code）
            key = (
                row["sheet_code"],
                row["report_type"],
                row["period_type"],
                row["node_name"],
                row["metric_category"],
            )
            dedup_key_map[key] = row

    final_report_rows = list(dedup_key_map.values())
    removed_count = len(all_report_rows) - len(final_report_rows)
    print(f"  去重后: {len(final_report_rows)} 条报表数据 (移除 {removed_count} 条重复)")

    # 解析 2.1 末尾附加的突围累计整体目标指标
    print("\n解析突围累计整体目标附加指标 (2.1 tail columns)...")
    tuwei_extra_rows = parse_tuwei_cumulative_extra_metrics(wb, valid_nodes)
    print(f"共 {len(tuwei_extra_rows)} 条突围累计附加指标数据")

    if tuwei_extra_rows:
        final_report_rows.extend(tuwei_extra_rows)
        print(f"  合并后报表数据: {len(final_report_rows)} 条")

    print("\n解析战略预算规划 (5)...")
    strategy_plan_rows = parse_strategy_budget_sheet(wb)
    print(f"共 {len(strategy_plan_rows)} 条战略预算规划数据")

    print("\n导入摘要:")
    print(f"  组织节点: {len(org_hierarchy_rows)}")
    print(f"  经营报表: {len(final_report_rows)}")
    print(f"  战略预算: {len(strategy_plan_rows)}")
    print("  目标表: edu_org_hierarchy, edu_biz_report, edu_strategy_budget_plan")

    if args.dry_run:
        print("\nDry run 完成，未写入数据库。")
        wb.close()
        return

    if not args.confirm:
        print("\n检测到将执行覆盖写入。请追加 --confirm 后重试。")
        wb.close()
        sys.exit(2)

    supabase = require_supabase_env()

    # 清空目标表
    print("\n清空目标表...")
    clear_table(supabase["SUPABASE_URL"], supabase["HEADERS"], "edu_biz_report")
    clear_table(supabase["SUPABASE_URL"], supabase["HEADERS"], "edu_org_hierarchy")
    clear_table(supabase["SUPABASE_URL"], supabase["HEADERS"], STRATEGY_PLAN_TABLE)
    print()

    # 写入 Supabase（先写入层级表，再写入数据表，但无外键约束）
    print("\n写入 Supabase...")

    if org_hierarchy_rows:
        n = upsert_batch(supabase["SUPABASE_URL"], supabase["HEADERS"], "edu_org_hierarchy", org_hierarchy_rows)
        print(f"  edu_org_hierarchy: 写入 {n}/{len(org_hierarchy_rows)} 条")

    if final_report_rows:
        n = upsert_batch(supabase["SUPABASE_URL"], supabase["HEADERS"], "edu_biz_report", final_report_rows)
        print(f"  edu_biz_report: 写入 {n}/{len(final_report_rows)} 条")

    if strategy_plan_rows:
        n = upsert_batch(supabase["SUPABASE_URL"], supabase["HEADERS"], STRATEGY_PLAN_TABLE, strategy_plan_rows)
        print(f"  {STRATEGY_PLAN_TABLE}: 写入 {n}/{len(strategy_plan_rows)} 条")

    print("\n完成!")
    wb.close()


if __name__ == "__main__":
    main()
